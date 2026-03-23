"""
Part C Evaluation Pipeline
Cross-verifies student Part C answers (recall questions) against their Part B submissions.

5 questions, students answer best 4 out of 5.
Each question worth 1.25 marks → total 5 marks.
Partial credit allowed based on coherence between Part C answer and Part B work.

Part C Questions map to Part B tasks:
  Q1 (24107) → Task 2.3 (Results & Reproducibility)
  Q2 (24108) → Task 3.1 (Ablation Study)
  Q3 (24109) → Task 3.2 (Failure Mode) + Task 1.2 (Key Assumptions)
  Q4 (24110) → Task 2.2 (Code Reproduction)
  Q5 (24111) → Task 4.1 (Report/Reflection)
"""

import json
import os
import sys
import time
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from agents.llm_client import call_llm_json
from agents.github_checker import parse_github_url, check_repo_exists, fetch_file_content
from agents.part_b_evaluator import fetch_notebook_content, resolve_notebook_name

OUTPUT_DIR = Path(__file__).parent.parent / "output"
SCORES_DIR = OUTPUT_DIR / "part_c_scores"
SCORES_DIR.mkdir(parents=True, exist_ok=True)

# Part C question IDs and their corresponding Part B tasks
PART_C_QUESTIONS = {
    24107: {
        "id": "Q1",
        "part_b_task": "task_2_3",
        "topic": "Results & Reproducibility",
        "description": "Recall exact metric value from Task 2.3 and explain gap with paper's value",
    },
    24108: {
        "id": "Q2",
        "part_b_task": "task_3_1",
        "topic": "Ablation Study",
        "description": "Name two ablated components from Task 3.1 and their performance impact",
    },
    24109: {
        "id": "Q3",
        "part_b_task": "task_3_2",
        "related_task": "task_1_2",
        "topic": "Failure Mode & Assumptions",
        "description": "Describe failure mode from Task 3.2 and connect to assumption from Task 1.2",
    },
    24110: {
        "id": "Q4",
        "part_b_task": "task_2_2",
        "topic": "Code Reproduction",
        "description": "Recall a code block from Task 2.2 and explain what it computes",
    },
    24111: {
        "id": "Q5",
        "part_b_task": "task_4_1",
        "topic": "Report/Reflection",
        "description": "Expand on surprising finding with specific result/plot/number",
    },
}

MAX_MARKS_PER_Q = 1.25
NUM_BEST = 4  # Take best 4 out of 5
TOTAL_MARKS = MAX_MARKS_PER_Q * NUM_BEST  # 5.0


def parse_part_c_from_excel(xlsx_path: str) -> tuple[dict, dict]:
    """
    Parse Part C answers and user_id → roll_number mapping from the Excel file.

    Returns:
        answers: {user_id: {question_id: answer_text, ...}}
        uid_to_roll: {user_id: roll_number}
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Build user_id → roll_number mapping from Manual Evaluation sheet
    uid_to_roll = {}
    uid_to_name = {}
    if "Manual Evaluation" in wb.sheetnames:
        ws_manual = wb["Manual Evaluation"]
        for i, row in enumerate(ws_manual.iter_rows(values_only=True)):
            if i <= 1:
                continue  # Skip header rows
            user_id = row[1]
            enrolment = row[2]
            name = row[3]
            if user_id and enrolment:
                uid_to_roll[int(float(user_id))] = str(int(float(enrolment)))
                uid_to_name[int(float(user_id))] = name

    # Parse Part C answers from Coding sheet
    part_c_qids = set(PART_C_QUESTIONS.keys())
    answers = {}  # {user_id: {qid: answer_text}}

    if "Coding" in wb.sheetnames:
        ws_coding = wb["Coding"]
        for i, row in enumerate(ws_coding.iter_rows(values_only=True)):
            if i == 0:
                continue
            user_id = row[0]
            answer_text = row[6]  # text_field column
            question_id = row[7]  # assignment_question_id column

            if not user_id or not question_id:
                continue

            uid = int(float(user_id))
            qid = int(float(question_id))

            if qid in part_c_qids and answer_text:
                if uid not in answers:
                    answers[uid] = {}
                answers[uid][qid] = str(answer_text).strip()

    wb.close()

    print(f"  Parsed Part C: {len(answers)} students with answers, {len(uid_to_roll)} user→roll mappings")
    return answers, uid_to_roll, uid_to_name


def fetch_part_b_content_for_task(owner: str, repo: str, task_name: str, branch: str = "main", file_map: dict = None) -> str:
    """
    Fetch the relevant Part B notebook content for a given task.
    Uses file_map for direct path lookup when available.
    Returns a string summary of the notebook cells.
    """
    # Use file_map if available (no guessing needed)
    if file_map and task_name in file_map.get("notebooks", {}):
        nb_path = file_map["notebooks"][task_name]
        nb_data = fetch_notebook_content(owner, repo, nb_path, branch)
    else:
        # Fallback: try resolve + alternates
        try:
            nb_path = resolve_notebook_name(owner, repo, task_name, branch)
        except Exception:
            nb_path = None

        if not nb_path:
            nb_path = f"partB/{task_name}.ipynb"

        nb_data = fetch_notebook_content(owner, repo, nb_path, branch)
        if nb_data.get("status") == "not_found":
            alternates = [
                f"Part_B/{task_name}.ipynb",
                f"part-B/{task_name}.ipynb",
                f"PartB/{task_name}.ipynb",
                f"{task_name}.ipynb",
                f"partb/{task_name}.ipynb",
            ]
            for alt in alternates:
                nb_data = fetch_notebook_content(owner, repo, alt, branch)
                if nb_data.get("status") != "not_found":
                    break

    if nb_data.get("status") == "not_found":
        return ""

    # Build a text summary from cells
    text_parts = []
    for cell in nb_data.get("cells", []):
        cell_type = cell.get("type", "")
        source = cell.get("source", "")
        if cell_type == "markdown":
            text_parts.append(f"[MARKDOWN]\n{source[:1000]}")
        elif cell_type == "code":
            text_parts.append(f"[CODE]\n{source[:500]}")

    combined = "\n\n".join(text_parts)
    # Truncate to fit LLM context
    return combined[:8000]


def evaluate_single_question(
    question_id: int,
    student_answer: str,
    part_b_content: str,
    question_info: dict,
) -> dict:
    """
    Evaluate a single Part C question by cross-verifying against Part B content.
    Returns: {"score": float, "reasoning": str, "coherence": str}
    """
    q_topic = question_info["topic"]
    q_desc = question_info["description"]
    q_id = question_info["id"]
    max_marks = MAX_MARKS_PER_Q

    if not part_b_content:
        # No Part B content found - can still evaluate the answer quality
        prompt = f"""You are evaluating a student's Part C answer for an ML exam.
Part C tests whether the student genuinely did their Part B work (not copy-pasted from AI).

Question topic: {q_topic}
Question asks student to: {q_desc}

Student's Part C answer:
\"\"\"{student_answer[:2000]}\"\"\"

NOTE: The student's Part B notebook for this task could not be found.
Evaluate the answer on its own merits - does it contain specific details
(exact numbers, method names, concrete observations) that suggest genuine work?
Or is it vague and generic, suggesting the student didn't actually do the work?

Return JSON:
{{
    "score": <float 0 to {max_marks}, with increments of 0.25>,
    "coherence": "high" | "medium" | "low" | "none",
    "reasoning": "<2-3 sentences explaining your evaluation>"
}}"""
    else:
        prompt = f"""You are evaluating a student's Part C answer by cross-verifying it against their actual Part B submission.
Part C tests whether the student genuinely did their Part B work themselves.

**Question topic:** {q_topic}
**What the question asks:** {q_desc}

**Student's Part C answer (written from memory):**
\"\"\"{student_answer[:2000]}\"\"\"

**Student's actual Part B submission ({question_info['part_b_task']}):**
\"\"\"{part_b_content[:6000]}\"\"\"

**Evaluate coherence between Part C answer and Part B submission:**
1. Do specific details (metric values, component names, observations) in Part C match Part B?
2. Does the student demonstrate genuine understanding of what they implemented?
3. Are there inconsistencies that suggest the student didn't do the Part B work themselves?
4. Does the Part C answer reference concrete, verifiable details from Part B?

**Scoring guide (out of {max_marks}):**
- {max_marks}: Perfect coherence. Specific details match exactly, clear genuine understanding.
- {max_marks * 0.75:.2f}: Good coherence. Most details match, minor inconsistencies.
- {max_marks * 0.5:.2f}: Partial coherence. Some details match but also vague or partially wrong.
- {max_marks * 0.25:.2f}: Low coherence. Mostly vague/generic, few matching details.
- 0: No coherence. Details contradict Part B or answer is completely generic.

Return JSON:
{{
    "score": <float 0 to {max_marks}, with increments of 0.25>,
    "coherence": "high" | "medium" | "low" | "none",
    "reasoning": "<2-3 sentences explaining the match/mismatch between Part C and Part B>"
}}"""

    system_prompt = (
        "You are a strict but fair exam evaluator for an Advanced ML course. "
        "Your job is to verify if students genuinely did their Part B work by "
        "cross-checking their Part C recall answers against their actual submissions. "
        "Be precise and objective. Always return valid JSON."
    )

    result = call_llm_json(prompt, system_prompt)

    if not result or "_raw_response" in result:
        return {
            "question_id": question_id,
            "q_id": q_id,
            "score": 0,
            "coherence": "error",
            "reasoning": "LLM evaluation failed",
        }

    return {
        "question_id": question_id,
        "q_id": q_id,
        "score": min(max_marks, max(0, float(result.get("score", 0)))),
        "coherence": result.get("coherence", "unknown"),
        "reasoning": result.get("reasoning", ""),
    }


def evaluate_student_part_c(
    roll_number: str,
    student_answers: dict,  # {qid: answer_text}
    github_url: str,
) -> dict:
    """
    Evaluate all Part C questions for a single student.
    Takes best 4 out of 5.

    Args:
        roll_number: Student's roll number
        student_answers: {question_id: answer_text}
        github_url: Student's GitHub repo URL

    Returns:
        dict with scores, reasoning, and total
    """
    print(f"\n  Part C evaluation for {roll_number}")

    # Parse GitHub URL and build file_map
    owner, repo = "", ""
    branch = "main"
    fm = None
    if github_url:
        owner, repo = parse_github_url(github_url)
        if owner and repo:
            repo_info = check_repo_exists(owner, repo)
            if repo_info.get("exists"):
                branch = repo_info.get("default_branch", "main")
                # Build file_map once for all questions
                from agents.github_checker import get_file_map
                fm = get_file_map(owner, repo, branch)
            else:
                print(f"    Repo not found: {github_url}")
                owner, repo = "", ""

    question_results = []

    for qid, answer_text in student_answers.items():
        if qid not in PART_C_QUESTIONS:
            continue

        q_info = PART_C_QUESTIONS[qid]
        task_name = q_info["part_b_task"]

        # Fetch corresponding Part B content
        part_b_content = ""
        if owner and repo:
            print(f"    Fetching Part B {task_name} for cross-verification...")
            part_b_content = fetch_part_b_content_for_task(owner, repo, task_name, branch, file_map=fm)

            # For Q3, also fetch task_1_2 if available
            if qid == 24109 and "related_task" in q_info:
                related_content = fetch_part_b_content_for_task(
                    owner, repo, q_info["related_task"], branch, file_map=fm
                )
                if related_content:
                    part_b_content += f"\n\n--- Related Task ({q_info['related_task']}) ---\n{related_content[:3000]}"

        if part_b_content:
            print(f"    Evaluating {q_info['id']}: {q_info['topic']} (with Part B cross-check)")
        else:
            print(f"    Evaluating {q_info['id']}: {q_info['topic']} (standalone - no Part B found)")

        result = evaluate_single_question(qid, answer_text, part_b_content, q_info)
        question_results.append(result)

        time.sleep(1)  # Rate limit buffer

    # Sort by score descending, take best 4
    question_results.sort(key=lambda x: x["score"], reverse=True)
    best_4 = question_results[:NUM_BEST]
    total_score = sum(r["score"] for r in best_4)

    result = {
        "roll_number": roll_number,
        "all_questions": question_results,
        "best_4": best_4,
        "total_score": round(total_score, 2),
        "max_possible": TOTAL_MARKS,
        "questions_answered": len(question_results),
        "questions_used": len(best_4),
    }

    # Save individual score file
    score_path = SCORES_DIR / f"{roll_number}_part_c.json"
    with open(score_path, "w") as f:
        json.dump(result, f, indent=2)

    print(f"    Part C total: {total_score:.2f}/{TOTAL_MARKS} (best {len(best_4)} of {len(question_results)} answered)")

    return result


def run_part_c_evaluation(
    xlsx_path: str,
    valid_students: list[dict],
    output_dir: Path = None,
) -> list[dict]:
    """
    Run Part C evaluation for all students.

    Args:
        xlsx_path: Path to the Advance ML Midsem Submissions - RU.xlsx file
        valid_students: List of valid student dicts from Phase 0 (with roll_number, github_url)
        output_dir: Optional output directory override

    Returns:
        List of Part C result dicts
    """
    global SCORES_DIR
    if output_dir:
        SCORES_DIR = output_dir / "part_c_scores"
        SCORES_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 70)
    print("PART C EVALUATION: Cross-Verification of Part B Work")
    print("=" * 70)

    # Parse Part C data from Excel
    answers, uid_to_roll, uid_to_name = parse_part_c_from_excel(xlsx_path)

    # Build roll → github_url mapping from valid students
    roll_to_github = {}
    for s in valid_students:
        roll = str(s.get("roll_number", ""))
        github = s.get("github_repo_link", "") or s.get("github_link", "")
        if roll and github:
            roll_to_github[roll] = github

    all_results = []
    evaluated = 0

    for uid, student_answers in answers.items():
        roll = uid_to_roll.get(uid, "")
        name = uid_to_name.get(uid, "Unknown")

        if not roll:
            print(f"  SKIP user_id={uid}: No roll number mapping found")
            continue

        github_url = roll_to_github.get(roll, "")

        try:
            result = evaluate_student_part_c(roll, student_answers, github_url)
            result["full_name"] = name
            all_results.append(result)
            evaluated += 1
        except Exception as e:
            print(f"  ERROR evaluating Part C for {roll}: {e}")
            all_results.append({
                "roll_number": roll,
                "full_name": name,
                "error": str(e),
                "total_score": 0,
                "max_possible": TOTAL_MARKS,
            })

    # Save aggregate results
    if output_dir:
        agg_path = output_dir / "part_c_all_results.json"
    else:
        agg_path = SCORES_DIR.parent / "part_c_all_results.json"

    with open(agg_path, "w") as f:
        json.dump(all_results, f, indent=2, default=str)

    print(f"\nPart C Evaluation Complete: {evaluated}/{len(answers)} students evaluated")
    avg = sum(r.get("total_score", 0) for r in all_results) / max(len(all_results), 1)
    print(f"Average Part C score: {avg:.2f}/{TOTAL_MARKS}")

    return all_results

"""
Phase 0: Data Cleanup & Deduplication

Reads Part A submissions, handles:
1. Multiple submissions per student (keep latest, count extras for penalty)
2. Duplicate paper selections (earliest timestamp wins)
3. Extracts clean student records
4. Validates basic fields
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / "evaluator" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def parse_part_a_submissions(xlsx_path: str) -> list[dict]:
    """Parse the Part A Excel file into a list of submission dicts."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    submissions = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        timestamp = row[0]
        if isinstance(timestamp, str):
            timestamp = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")

        sub = {
            "timestamp": timestamp.isoformat() if isinstance(timestamp, datetime) else str(timestamp),
            "timestamp_dt": timestamp,
            "full_name": str(row[1]).strip() if row[1] else "",
            "authors": str(row[2]).strip() if row[2] else "",
            "year_of_publication": int(row[3]) if row[3] else None,
            "paper_link": str(row[4]).strip() if row[4] else "",
            "primary_method": str(row[5]).strip() if row[5] else "",
            "roll_number": str(int(row[6])) if row[6] else "",
            "paper_title": str(row[7]).strip() if row[7] else "",
            "email": str(row[8]).strip() if row[8] else "",
            "venue": str(row[9]).strip() if row[9] else "",
            "is_methodological": str(row[10]).strip() if row[10] else "",
            "publicly_accessible_data": str(row[11]).strip() if row[11] else "",
            "requires_special": str(row[12]).strip() if row[12] else "",
            "understands_responsibility": str(row[13]).strip() if row[13] else "",
            "checked_response_sheet": str(row[14]).strip() if row[14] else "",
            "confirms_duplicate_rule": str(row[15]).strip() if row[15] else "",
            "why_chosen": str(row[16]).strip() if row[16] else "",
            "email_address_2": str(row[17]).strip() if row[17] else "",
            "github_repo": str(row[18]).strip() if row[18] else "",
        }
        submissions.append(sub)

    print(f"Parsed {len(submissions)} raw submissions from Excel")
    return submissions


def handle_resubmissions(submissions: list[dict]) -> tuple[list[dict], dict]:
    """
    Group by roll number. Keep the LATEST submission per student.
    Track submission count for penalty (20% per extra beyond first).
    """
    by_roll = {}
    for sub in submissions:
        roll = sub["roll_number"]
        if roll not in by_roll:
            by_roll[roll] = []
        by_roll[roll].append(sub)

    final_submissions = []
    resubmission_penalties = {}

    for roll, subs in by_roll.items():
        subs_sorted = sorted(subs, key=lambda x: x["timestamp_dt"])
        latest = subs_sorted[-1]
        submission_count = len(subs_sorted)

        if submission_count > 1:
            extra = submission_count - 1
            penalty_pct = extra * 20
            resubmission_penalties[roll] = {
                "student_name": latest["full_name"],
                "total_submissions": submission_count,
                "extra_submissions": extra,
                "penalty_percentage": penalty_pct,
                "timestamps": [s["timestamp"] for s in subs_sorted],
            }

        final_submissions.append(latest)

    print(f"\nResubmission Analysis:")
    print(f"  Unique students: {len(final_submissions)}")
    print(f"  Students with resubmissions: {len(resubmission_penalties)}")
    for roll, info in resubmission_penalties.items():
        print(f"    {roll} ({info['student_name']}): {info['total_submissions']} submissions -> {info['penalty_percentage']}% penalty")

    return final_submissions, resubmission_penalties


def detect_duplicate_papers(submissions: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Detect duplicate paper selections.
    First-come-first-serve: earliest timestamp keeps the paper.
    All later duplicates are disqualified from Parts B & C.
    """
    # Normalize paper titles for comparison
    def normalize_title(title):
        return title.lower().strip().rstrip(".")

    by_paper = {}
    for sub in submissions:
        norm_title = normalize_title(sub["paper_title"])
        if norm_title not in by_paper:
            by_paper[norm_title] = []
        by_paper[norm_title].append(sub)

    valid = []
    disqualified = []

    for norm_title, subs in by_paper.items():
        subs_sorted = sorted(subs, key=lambda x: x["timestamp_dt"])
        # First submission keeps the paper
        valid.append(subs_sorted[0])
        # All later duplicates are disqualified
        for dup in subs_sorted[1:]:
            dup["disqualified_reason"] = f"Duplicate paper (first selected by {subs_sorted[0]['roll_number']} at {subs_sorted[0]['timestamp']})"
            disqualified.append(dup)

    print(f"\nDuplicate Paper Analysis:")
    print(f"  Unique papers: {len(by_paper)}")
    print(f"  Valid (unique) submissions: {len(valid)}")
    print(f"  Disqualified (duplicates): {len(disqualified)}")
    if disqualified:
        for d in disqualified:
            print(f"    {d['roll_number']} ({d['full_name']}): \"{d['paper_title'][:60]}...\" -> {d['disqualified_reason']}")

    return valid, disqualified


def validate_basic_fields(submissions: list[dict]) -> list[dict]:
    """Add validation flags to each submission."""
    from config.core_a_star_venues import CORE_A_STAR_VENUES, VALID_YEARS

    for sub in submissions:
        issues = []

        # Year check
        if sub["year_of_publication"] and sub["year_of_publication"] not in VALID_YEARS:
            issues.append(f"Year {sub['year_of_publication']} not in 2009-2012")

        # Venue check (basic string matching)
        venue_lower = sub["venue"].lower().strip()
        venue_match = any(v in venue_lower or venue_lower in v for v in CORE_A_STAR_VENUES)
        if not venue_match:
            issues.append(f"Venue '{sub['venue']}' not recognized as CORE A*")

        # Method check
        method_lower = sub["primary_method"].lower()
        method_valid = any(m in method_lower for m in ["arima", "time series", "gmm", "gaussian mixture", "svm", "support vector"])
        if not method_valid:
            issues.append(f"Method '{sub['primary_method']}' may not align with ARIMA/GMM/SVM")

        # GitHub repo check
        if not sub["github_repo"] or "github.com" not in sub["github_repo"].lower():
            issues.append("Missing or invalid GitHub repo URL")

        # Paper link check
        if not sub["paper_link"]:
            issues.append("Missing paper link")

        sub["validation_issues"] = issues
        sub["has_issues"] = len(issues) > 0

    issues_count = sum(1 for s in submissions if s["has_issues"])
    print(f"\nBasic Validation:")
    print(f"  Submissions with issues: {issues_count}/{len(submissions)}")
    for sub in submissions:
        if sub["has_issues"]:
            print(f"    {sub['roll_number']} ({sub['full_name']}): {sub['validation_issues']}")

    return submissions


def parse_part_b_submissions(csv_path: str) -> list[dict]:
    """Parse Part B CSV submissions."""
    import csv

    submissions = []
    with open(csv_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sub = {
                "timestamp": row.get("Timestamp", ""),
                "email": row.get("Email Address", ""),
                "full_name": row.get("Full Name", ""),
                "roll_number": str(row.get("Roll Number / University ID", "")).strip(),
                "email_institute": row.get("Email ID (Institute email preferred)", ""),
                "paper_title": row.get("Paper Title You Submitted in Part A", ""),
                "github_repo": row.get("Github Repository link with Report, Code and LLM Usage (JSON file) [Use the Same Repo which you used for Part A.", ""),
            }
            submissions.append(sub)

    print(f"\nPart B Submissions: {len(submissions)}")
    for s in submissions:
        print(f"  {s['roll_number']} ({s['full_name']}): {s['paper_title'][:60]}")

    return submissions


def generate_summary(valid, disqualified, penalties, part_b, output_dir=None):
    """Generate a summary JSON with all processed data."""
    target_dir = output_dir or OUTPUT_DIR

    # Remove non-serializable datetime objects
    def clean_sub(sub):
        cleaned = {k: v for k, v in sub.items() if k != "timestamp_dt"}
        return cleaned

    summary = {
        "generated_at": datetime.now().isoformat(),
        "stats": {
            "total_valid_part_a": len(valid),
            "total_disqualified": len(disqualified),
            "students_with_resubmission_penalty": len(penalties),
            "total_part_b_submissions": len(part_b),
        },
        "resubmission_penalties": penalties,
        "disqualified_students": [clean_sub(d) for d in disqualified],
        "valid_students": [clean_sub(v) for v in valid],
        "part_b_students": part_b,
        "unique_papers": list(set(s["paper_title"] for s in valid)),
    }

    target_dir.mkdir(parents=True, exist_ok=True)
    output_path = target_dir / "phase0_summary.json"
    with open(output_path, "w") as f:
        json.dump(summary, f, indent=2, default=str)

    print(f"\nSummary written to {output_path}")
    return summary


def run_phase0_web(xlsx_path: str, csv_path: str, output_dir: Path) -> dict:
    """Web-compatible Phase 0 that accepts arbitrary file paths and output directory."""
    print("=" * 70)
    print("PHASE 0: DATA CLEANUP & DEDUPLICATION (Web)")
    print("=" * 70)

    raw_submissions = parse_part_a_submissions(str(xlsx_path))
    unique_submissions, penalties = handle_resubmissions(raw_submissions)
    valid, disqualified = detect_duplicate_papers(unique_submissions)

    import importlib
    sys.path.insert(0, str(Path(__file__).parent))
    valid = validate_basic_fields(valid)
    part_b = parse_part_b_submissions(str(csv_path))

    summary = generate_summary(valid, disqualified, penalties, part_b, output_dir)
    return summary


def main():
    xlsx_path = BASE_DIR / "Advance ML Midsem Part A Submission Form (Responses).xlsx"
    csv_path = BASE_DIR / "Advance ML Midsem Part B Submission Form (Responses) - Form Responses 1.csv"

    print("=" * 70)
    print("PHASE 0: DATA CLEANUP & DEDUPLICATION")
    print("=" * 70)

    # Step 1: Parse raw submissions
    raw_submissions = parse_part_a_submissions(str(xlsx_path))

    # Step 2: Handle resubmissions (keep latest per student)
    unique_submissions, penalties = handle_resubmissions(raw_submissions)

    # Step 3: Detect duplicate papers
    valid, disqualified = detect_duplicate_papers(unique_submissions)

    # Step 4: Basic field validation
    valid = validate_basic_fields(valid)

    # Step 5: Parse Part B
    part_b = parse_part_b_submissions(str(csv_path))

    # Step 6: Cross-reference Part B students with Part A
    print("\n" + "=" * 70)
    print("CROSS-REFERENCE: Part B students in Part A")
    print("=" * 70)
    valid_rolls = {s["roll_number"] for s in valid}
    disq_rolls = {s["roll_number"] for s in disqualified}
    for b in part_b:
        roll = b["roll_number"]
        if roll in valid_rolls:
            print(f"  {roll} ({b['full_name']}): VALID Part A submission")
        elif roll in disq_rolls:
            print(f"  {roll} ({b['full_name']}): DISQUALIFIED (duplicate paper)")
        else:
            print(f"  {roll} ({b['full_name']}): NO Part A submission found!")

    # Step 7: Generate summary
    summary = generate_summary(valid, disqualified, penalties, part_b)

    print("\n" + "=" * 70)
    print("PHASE 0 COMPLETE")
    print("=" * 70)
    print(f"  Valid students for Part A evaluation: {len(valid)}")
    print(f"  Disqualified students: {len(disqualified)}")
    print(f"  Part B students to evaluate: {len(part_b)}")
    print(f"  Unique papers to analyze: {len(summary['unique_papers'])}")

    return summary


if __name__ == "__main__":
    os.chdir(str(BASE_DIR / "evaluator"))
    summary = main()
